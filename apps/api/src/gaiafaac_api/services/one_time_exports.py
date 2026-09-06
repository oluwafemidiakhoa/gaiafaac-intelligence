from __future__ import annotations

import io
import json
import re
from collections.abc import Iterator
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_TEAL = "07594F"
_DARK_TEAL = "041915"
_AMBER = "F7C948"
_LIGHT_TEAL = "E8F5F2"
_LIGHT_AMBER = "FFF6D8"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MIME = "application/pdf"


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return str(value)


def _excel_value(value: object) -> object:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    text = _display(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _flatten_mapping(value: dict[str, Any], *, prefix: str = "") -> dict[str, object]:
    flattened: dict[str, object] = {}
    for key, item in value.items():
        path = f"{prefix}.{key}" if prefix else key
        if isinstance(item, dict):
            flattened.update(_flatten_mapping(item, prefix=path))
        elif isinstance(item, list):
            flattened[path] = _display(item)
        else:
            flattened[path] = item
    return flattened


def _scalar_rows(value: object, *, path: str = "artifact") -> Iterator[tuple[str, object]]:
    if isinstance(value, dict):
        for key, item in value.items():
            yield from _scalar_rows(item, path=f"{path}.{key}")
    elif isinstance(value, list):
        return
    else:
        yield path, value


def _table_rows(
    value: object,
    *,
    path: str = "artifact",
) -> Iterator[tuple[str, list[dict[str, Any]]]]:
    if isinstance(value, dict):
        for key, item in value.items():
            yield from _table_rows(item, path=f"{path}.{key}")
        return
    if not isinstance(value, list) or not value:
        return
    if all(isinstance(item, dict) for item in value):
        yield path, value
        for index, item in enumerate(value):
            yield from _table_rows(item, path=f"{path}.{index}")


def _sheet_title(path: str, used: set[str]) -> str:
    last = path.rsplit(".", maxsplit=1)[-1].replace("_", " ").strip().title() or "Evidence"
    if last == "Rows":
        last = "Evidence Rows"
    base = re.sub(r"[\\/*?:\[\]]", "-", last)[:31]
    title = base
    suffix = 2
    while title in used:
        marker = f" {suffix}"
        title = f"{base[: 31 - len(marker)]}{marker}"
        suffix += 1
    used.add(title)
    return title


def _filename(product_code: str, purchase_id: str, suffix: str) -> str:
    product = re.sub(r"[^a-z0-9]+", "-", product_code.casefold()).strip("-") or "project"
    short_id = re.sub(r"[^a-zA-Z0-9]", "", purchase_id)[:12] or "order"
    return f"gaia-{product}-{short_id}.{suffix}"


def _style_summary_sheet(sheet, rows: list[tuple[str, object]]) -> None:
    sheet.merge_cells("A1:D1")
    title = sheet["A1"]
    title.value = "Gaia Fiscal Intelligence"
    title.font = Font(size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=_DARK_TEAL)
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells("A2:D2")
    subtitle = sheet["A2"]
    subtitle.value = "Governed project deliverable — frozen to the paid order"
    subtitle.font = Font(size=11, bold=True, color=_TEAL)
    subtitle.fill = PatternFill("solid", fgColor=_LIGHT_TEAL)

    sheet.append([])
    sheet.append(["Field", "Value"])
    for cell in sheet[4]:
        cell.font = Font(bold=True, color=_DARK_TEAL)
        cell.fill = PatternFill("solid", fgColor=_AMBER)

    for key, value in rows:
        sheet.append([key, _excel_value(value)])

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 96
    for row in sheet.iter_rows(min_row=5):
        row[0].font = Font(bold=True, color=_TEAL)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A5"


def _style_table_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    flattened = [_flatten_mapping(row) for row in rows]
    headers: list[str] = []
    for row in flattened:
        for header in row:
            if header not in headers:
                headers.append(header)
    if not headers:
        headers = ["record"]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_TEAL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in flattened:
        sheet.append([_excel_value(row.get(header)) for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, header in enumerate(headers, start=1):
        max_width = len(header)
        for cell in sheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=sheet.max_row,
        ):
            for item in cell:
                text = _display(item.value)
                max_width = max(max_width, min(len(text), 72))
                item.alignment = Alignment(vertical="top", wrap_text=True)
                if header.endswith("source_url") and text.startswith(("https://", "http://")):
                    item.hyperlink = text
                    item.font = Font(color="0563C1", underline="single")
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_width + 2, 12), 56
        )


def build_one_time_excel(
    *,
    purchase_id: str,
    product_code: str,
    amount_naira: str,
    currency: str,
    completed_at: object,
    artifact: dict[str, Any],
) -> tuple[str, str, bytes]:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary_rows: list[tuple[str, object]] = [
        ("Order ID", purchase_id),
        ("Product", product_code.replace("_", " ").title()),
        ("Product code", product_code),
        ("Amount paid", amount_naira),
        ("Currency", currency),
        ("Payment completed", completed_at),
        ("Artifact schema", artifact.get("schema")),
        ("Evidence captured", artifact.get("captured_at")),
    ]
    request = artifact.get("request")
    if isinstance(request, dict):
        for key, value in _flatten_mapping(request, prefix="request").items():
            summary_rows.append((key, value))

    excluded = {"schema", "captured_at", "request"}
    for path, value in _scalar_rows(
        {key: value for key, value in artifact.items() if key not in excluded}
    ):
        summary_rows.append((path, value))

    _style_summary_sheet(summary, summary_rows)

    used = {"Summary"}
    tables = list(_table_rows(artifact))
    for path, rows in tables:
        sheet = workbook.create_sheet(_sheet_title(path, used))
        _style_table_sheet(sheet, rows)

    if not tables:
        sheet = workbook.create_sheet("Artifact")
        sheet.append(["Path", "Value"])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_TEAL)
        for path, value in _scalar_rows(artifact):
            sheet.append([path, _excel_value(value)])
        sheet.column_dimensions["A"].width = 48
        sheet.column_dimensions["B"].width = 100
        sheet.freeze_panes = "A2"

    workbook.properties.title = "Gaia Fiscal Intelligence governed project deliverable"
    workbook.properties.subject = product_code
    workbook.properties.creator = "Gaia Fiscal Intelligence"
    workbook.properties.description = (
        "Governed evidence deliverable frozen to a paid one-time Gaia project order."
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return _filename(product_code, purchase_id, "xlsx"), _XLSX_MIME, buffer.getvalue()


def _pdf_text(value: object) -> str:
    text = _display(value)
    if len(text) > 6_000:
        text = f"{text[:6_000]}…"
    return escape(text).replace("\n", "<br/>")


def build_one_time_pdf(
    *,
    purchase_id: str,
    product_code: str,
    amount_naira: str,
    currency: str,
    completed_at: object,
    artifact: dict[str, Any],
) -> tuple[str, str, bytes]:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="Gaia Fiscal Intelligence governed project deliverable",
        author="Gaia Fiscal Intelligence",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GaiaTitle",
        parent=styles["Title"],
        textColor=colors.HexColor(f"#{_DARK_TEAL}"),
        fontSize=24,
        leading=28,
        spaceAfter=4 * mm,
    )
    section_style = ParagraphStyle(
        "GaiaSection",
        parent=styles["Heading2"],
        textColor=colors.HexColor(f"#{_TEAL}"),
        fontSize=13,
        leading=16,
        spaceBefore=4 * mm,
        spaceAfter=2 * mm,
    )
    body_style = ParagraphStyle(
        "GaiaBody",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#263633"),
    )
    small_style = ParagraphStyle(
        "GaiaSmall",
        parent=body_style,
        fontSize=7,
        leading=9,
    )

    story: list[object] = [
        Paragraph("Gaia Fiscal Intelligence", title_style),
        Paragraph("Governed project deliverable — frozen to the paid order", section_style),
    ]

    order_rows = [
        ["Order ID", purchase_id],
        ["Product", product_code.replace("_", " ").title()],
        ["Amount paid", f"{currency} {amount_naira}"],
        ["Payment completed", _display(completed_at)],
        ["Artifact schema", _display(artifact.get("schema"))],
        ["Evidence captured", _display(artifact.get("captured_at"))],
    ]
    order_table = Table(
        [
            [
                Paragraph(f"<b>{escape(label)}</b>", body_style),
                Paragraph(_pdf_text(value), body_style),
            ]
            for label, value in order_rows
        ],
        colWidths=[48 * mm, 190 * mm],
    )
    order_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_LIGHT_TEAL}")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor(f"#{_TEAL}")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9D7D3")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend([order_table, Spacer(1, 3 * mm)])

    request = artifact.get("request")
    if isinstance(request, dict):
        story.append(Paragraph("Evidence boundary", section_style))
        request_rows = [
            [
                Paragraph(f"<b>{escape(key)}</b>", body_style),
                Paragraph(_pdf_text(value), body_style),
            ]
            for key, value in _flatten_mapping(request).items()
        ]
        request_table = Table(request_rows, colWidths=[55 * mm, 183 * mm])
        request_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_LIGHT_AMBER}")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDD3AE")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend([request_table, Spacer(1, 2 * mm)])

    tables = list(_table_rows(artifact))
    for path, rows in tables:
        heading = path.removeprefix("artifact.").replace("_", " ").replace(".", " / ").title()
        story.append(Paragraph(escape(heading), section_style))
        for index, row in enumerate(rows, start=1):
            story.append(Paragraph(f"Record {index}", small_style))
            flat = _flatten_mapping(row)
            record_rows = [
                [
                    Paragraph(f"<b>{escape(key)}</b>", small_style),
                    Paragraph(_pdf_text(value), small_style),
                ]
                for key, value in flat.items()
            ]
            if not record_rows:
                record_rows = [[Paragraph("Record", small_style), Paragraph("", small_style)]]
            table = Table(record_rows, colWidths=[64 * mm, 174 * mm], repeatRows=0)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F7F6")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#D5E0DD")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            story.extend([table, Spacer(1, 2 * mm)])

    statement = artifact.get("statement")
    if statement:
        story.extend(
            [
                Paragraph("Use statement", section_style),
                Paragraph(_pdf_text(statement), body_style),
            ]
        )

    story.extend(
        [
            Spacer(1, 5 * mm),
            Paragraph(
                "Integrity note: this document renders the governed artifact frozen to this paid order. "
                "Source organizations, source URLs and SHA-256 fingerprints remain part of the evidence record where available.",
                small_style,
            ),
        ]
    )

    def footer(canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#65726F"))
        canvas.drawString(14 * mm, 8 * mm, f"Gaia order {purchase_id}")
        canvas.drawRightString(283 * mm, 8 * mm, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return _filename(product_code, purchase_id, "pdf"), _PDF_MIME, buffer.getvalue()
