from __future__ import annotations

import io
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from gaiafaac_api.services.document_branding import (
    BRAND_NAME,
    BRAND_SUBTITLE,
    SAMPLE_NOTICE,
    brand_workbook,
    draw_pdf_branding,
)
from gaiafaac_api.services.one_time_exports import (
    _DARK_TEAL,
    _LIGHT_AMBER,
    _LIGHT_TEAL,
    _PDF_MIME,
    _TEAL,
    _XLSX_MIME,
    _display,
    _flatten_mapping,
    _pdf_text,
    _table_rows,
)
from gaiafaac_api.services.one_time_exports import (
    build_one_time_excel as _build_legacy_excel,
)


def _filename(product_code: str, purchase_id: str, suffix: str, *, sample: bool) -> str:
    product = product_code.replace("_", "-")
    if sample:
        return f"gaia-fiscal-intelligence-sample-{product}.{suffix}"
    short_id = "".join(character for character in purchase_id if character.isalnum())[:12]
    return f"gaia-fiscal-intelligence-{product}-{short_id or 'order'}.{suffix}"


def build_one_time_excel(
    *,
    purchase_id: str,
    product_code: str,
    amount_naira: str,
    currency: str,
    completed_at: object,
    artifact: dict[str, Any],
    sample: bool = False,
    jurisdiction: str | None = None,
) -> tuple[str, str, bytes]:
    _legacy_filename, _media_type, body = _build_legacy_excel(
        purchase_id=purchase_id,
        product_code=product_code,
        amount_naira=amount_naira,
        currency=currency,
        completed_at=completed_at,
        artifact=artifact,
    )
    workbook = load_workbook(io.BytesIO(body))
    summary = workbook["Summary"]
    summary["A1"] = BRAND_NAME
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor=_DARK_TEAL)
    summary["A2"] = BRAND_SUBTITLE
    if sample:
        summary["A3"] = SAMPLE_NOTICE
        summary["A3"].font = Font(size=10, bold=True, color="9C2A1B")

    brand_workbook(
        workbook,
        sample=sample,
        order_id=None if sample else purchase_id,
        jurisdiction=jurisdiction,
        generated_at=artifact.get("captured_at"),
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return (
        _filename(product_code, purchase_id, "xlsx", sample=sample),
        _XLSX_MIME,
        buffer.getvalue(),
    )


def build_one_time_pdf(
    *,
    purchase_id: str,
    product_code: str,
    amount_naira: str,
    currency: str,
    completed_at: object,
    artifact: dict[str, Any],
    sample: bool = False,
    jurisdiction: str | None = None,
) -> tuple[str, str, bytes]:
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"{BRAND_NAME} governed public-finance deliverable",
        author=BRAND_NAME,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GaiaFiscalTitle",
        parent=styles["Title"],
        textColor=colors.HexColor(f"#{_DARK_TEAL}"),
        fontSize=24,
        leading=28,
        spaceAfter=2 * mm,
    )
    subtitle_style = ParagraphStyle(
        "GaiaFiscalSubtitle",
        parent=styles["Heading2"],
        textColor=colors.HexColor(f"#{_TEAL}"),
        fontSize=13,
        leading=16,
        spaceAfter=3 * mm,
    )
    section_style = ParagraphStyle(
        "GaiaFiscalSection",
        parent=styles["Heading2"],
        textColor=colors.HexColor(f"#{_TEAL}"),
        fontSize=12,
        leading=15,
        spaceBefore=4 * mm,
        spaceAfter=2 * mm,
    )
    body_style = ParagraphStyle(
        "GaiaFiscalBody",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#263633"),
    )
    small_style = ParagraphStyle(
        "GaiaFiscalSmall",
        parent=body_style,
        fontSize=7,
        leading=9,
    )
    notice_style = ParagraphStyle(
        "GaiaFiscalSampleNotice",
        parent=body_style,
        textColor=colors.HexColor("#9C2A1B"),
        fontSize=10,
        leading=13,
        spaceAfter=3 * mm,
    )

    story: list[object] = [
        Paragraph(BRAND_NAME, title_style),
        Paragraph(BRAND_SUBTITLE, subtitle_style),
    ]
    if sample:
        story.append(Paragraph(f"<b>{escape(SAMPLE_NOTICE)}</b>", notice_style))

    order_rows = [
        ["Sample reference" if sample else "Order ID", purchase_id],
        ["Product", product_code.replace("_", " ").title()],
        ["Illustrative package value" if sample else "Amount paid", f"{currency} {amount_naira}"],
        ["Payment", "Not applicable — demonstration sample" if sample else _display(completed_at)],
        ["Artifact schema", _display(artifact.get("schema"))],
        ["Evidence captured", _display(artifact.get("captured_at"))],
    ]
    order_table = Table(
        [
            [
                Paragraph(f"<b>{escape(label)}</b>", body_style),
                Paragraph(_pdf_text(value), body_style),
            ]
            for label, value in order_rows
        ],
        colWidths=[54 * mm, 184 * mm],
    )
    order_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_LIGHT_TEAL}")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor(f"#{_TEAL}")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9D7D3")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.extend([order_table, Spacer(1, 3 * mm)])

    request = artifact.get("request")
    if isinstance(request, dict):
        story.append(Paragraph("Evidence boundary", section_style))
        request_rows = [
            [
                Paragraph(f"<b>{escape(key)}</b>", body_style),
                Paragraph(_pdf_text(value), body_style),
            ]
            for key, value in _flatten_mapping(request).items()
        ]
        request_table = Table(request_rows, colWidths=[55 * mm, 183 * mm])
        request_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_LIGHT_AMBER}")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDD3AE")),
                ]
            )
        )
        story.extend([request_table, Spacer(1, 2 * mm)])

    for path, rows in _table_rows(artifact):
        heading = path.removeprefix("artifact.").replace("_", " ").replace(".", " / ").title()
        story.append(Paragraph(escape(heading), section_style))
        for index, row in enumerate(rows, start=1):
            story.append(Paragraph(f"Record {index}", small_style))
            flat = _flatten_mapping(row)
            record_rows = [
                [
                    Paragraph(f"<b>{escape(key)}</b>", small_style),
                    Paragraph(_pdf_text(value), small_style),
                ]
                for key, value in flat.items()
            ] or [[Paragraph("Record", small_style), Paragraph("", small_style)]]
            table = Table(record_rows, colWidths=[64 * mm, 174 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F7F6")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#D5E0DD")),
                    ]
                )
            )
            story.extend([table, Spacer(1, 2 * mm)])

    statement = artifact.get("statement")
    if statement:
        story.extend(
            [
                Paragraph("Use statement", section_style),
                Paragraph(_pdf_text(statement), body_style),
            ]
        )

    story.extend(
        [
            Spacer(1, 5 * mm),
            Paragraph(
                "Integrity note: this document renders the governed artifact frozen to this order or demonstration capture. Source organizations, source URLs and SHA-256 fingerprints remain part of the evidence record where available.",
                small_style,
            ),
        ]
    )

    def page_brand(canvas, doc) -> None:
        draw_pdf_branding(
            canvas,
            doc,
            sample=sample,
            order_id=None if sample else purchase_id,
            jurisdiction=jurisdiction,
            generated_at=artifact.get("captured_at"),
        )

    document.build(story, onFirstPage=page_brand, onLaterPages=page_brand)
    return (
        _filename(product_code, purchase_id, "pdf", sample=sample),
        _PDF_MIME,
        buffer.getvalue(),
    )
