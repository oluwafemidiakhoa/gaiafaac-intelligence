from __future__ import annotations

from contextlib import suppress
from datetime import UTC, datetime
from hashlib import sha256

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.units import mm

BRAND_NAME = "GAIA FISCAL INTELLIGENCE"
BRAND_SUBTITLE = "Governed Public-Finance Intelligence"
PAID_WATERMARK = BRAND_NAME
SAMPLE_WATERMARK = "SAMPLE — GAIA FISCAL INTELLIGENCE — NOT FOR RESALE"
SAMPLE_NOTICE = "SAMPLE / DEMONSTRATION ONLY / NOT FOR RELIANCE IN FORMAL DECISION-MAKING"
_DARK_TEAL = "041915"
_TEAL = "07594F"
_LIGHT_TEAL = "E8F5F2"
_AMBER = "F7C948"
_SAMPLE_RED = "9C2A1B"


def _generated_label(value: object | None) -> str:
    if value in (None, ""):
        return datetime.now(UTC).isoformat()
    return str(value)


def document_fingerprint(
    *,
    sample: bool,
    order_id: str | None,
    jurisdiction: str | None,
    generated_at: object | None,
    artifact_sha256: str | None = None,
) -> str:
    generated = _generated_label(generated_at)
    material = "|".join(
        [
            BRAND_NAME,
            "sample" if sample else "paid",
            order_id or "",
            jurisdiction or "",
            generated,
            artifact_sha256 or "",
        ]
    )
    digest = sha256(material.encode("utf-8")).hexdigest().upper()
    return f"GFI-{digest[:4]}-{digest[4:8]}-{digest[8:12]}"


def _document_control_sheet(
    workbook: Workbook,
    *,
    sample: bool,
    order_id: str | None,
    jurisdiction: str | None,
    generated: str,
    fingerprint: str,
    artifact_sha256: str | None,
    verification_url: str | None,
) -> None:
    if "Document Control" in workbook.sheetnames:
        sheet = workbook["Document Control"]
    else:
        sheet = workbook.create_sheet("Document Control", 0)

    sheet.sheet_properties.tabColor = _TEAL
    sheet.merge_cells("A1:D1")
    sheet["A1"] = BRAND_NAME
    sheet["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=_DARK_TEAL)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells("A2:D2")
    sheet["A2"] = BRAND_SUBTITLE
    sheet["A2"].font = Font(size=11, bold=True, color=_TEAL)
    sheet["A2"].fill = PatternFill("solid", fgColor=_LIGHT_TEAL)

    row = 4
    if sample:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        notice = sheet.cell(row=row, column=1)
        notice.value = SAMPLE_NOTICE
        notice.font = Font(size=11, bold=True, color="FFFFFF")
        notice.fill = PatternFill("solid", fgColor=_SAMPLE_RED)
        notice.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 30
        row += 2

    sheet.cell(row=row, column=1, value="Document control")
    sheet.cell(row=row, column=2, value="Value")
    for cell in sheet[row][:2]:
        cell.font = Font(bold=True, color=_DARK_TEAL)
        cell.fill = PatternFill("solid", fgColor=_AMBER)
    row += 1

    entries = [
        ("Document class", "DEMONSTRATION SAMPLE" if sample else "PAID CUSTOMER DELIVERABLE"),
        ("Document ID", fingerprint),
        ("Artifact SHA-256", artifact_sha256 or "Not recorded"),
        ("Jurisdiction / scope", jurisdiction or "Governed evidence boundary"),
        ("Order ID", "Not applicable" if sample else (order_id or "Unavailable")),
        ("Generated", generated),
        ("Verification", verification_url or "Sample document — no paid receipt verification"),
        (
            "Commercial basis",
            "Customer pays for governed fiscal intelligence and evidence scope; PDF, Excel and JSON are included delivery formats.",
        ),
    ]
    for label, value in entries:
        sheet.cell(row=row, column=1, value=label)
        value_cell = sheet.cell(row=row, column=2, value=value)
        sheet.cell(row=row, column=1).font = Font(bold=True, color=_TEAL)
        sheet.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)
        if label == "Verification" and verification_url:
            value_cell.hyperlink = verification_url
            value_cell.style = "Hyperlink"
        row += 1

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 84
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 14
    sheet.freeze_panes = "A4"


def brand_workbook(
    workbook: Workbook,
    *,
    sample: bool = False,
    order_id: str | None = None,
    jurisdiction: str | None = None,
    generated_at: object | None = None,
    artifact_sha256: str | None = None,
    verification_url: str | None = None,
) -> None:
    """Apply institutional branding, artifact integrity and document control."""

    generated = _generated_label(generated_at)
    fingerprint = document_fingerprint(
        sample=sample,
        order_id=order_id,
        jurisdiction=jurisdiction,
        generated_at=generated,
        artifact_sha256=artifact_sha256,
    )
    watermark = SAMPLE_WATERMARK if sample else PAID_WATERMARK

    workbook.properties.title = f"{BRAND_NAME} governed public-finance deliverable"
    workbook.properties.creator = BRAND_NAME
    workbook.properties.keywords = f"{BRAND_NAME}; {fingerprint}; governed fiscal evidence"
    workbook.properties.description = (
        f"{SAMPLE_NOTICE}. Document ID {fingerprint}."
        if sample
        else f"Governed public-finance intelligence deliverable. Document ID {fingerprint}."
    )

    _document_control_sheet(
        workbook,
        sample=sample,
        order_id=order_id,
        jurisdiction=jurisdiction,
        generated=generated,
        fingerprint=fingerprint,
        artifact_sha256=artifact_sha256,
        verification_url=verification_url,
    )

    for sheet in workbook.worksheets:
        sheet.oddHeader.left.text = BRAND_NAME
        sheet.oddHeader.center.text = f"{watermark} · {fingerprint}"
        sheet.oddHeader.center.size = 12 if sample else 9
        sheet.oddHeader.center.font = "Arial,Bold"
        sheet.oddHeader.right.text = "SAMPLE" if sample else "GOVERNED EVIDENCE"
        sheet.evenHeader.left.text = BRAND_NAME
        sheet.evenHeader.center.text = f"{watermark} · {fingerprint}"
        sheet.evenHeader.right.text = "SAMPLE" if sample else "GOVERNED EVIDENCE"

        sheet.oddFooter.left.text = BRAND_SUBTITLE
        sheet.oddFooter.center.text = jurisdiction or "Governed evidence boundary"
        sheet.oddFooter.right.text = f"{fingerprint} · {generated} · Page &P of &N"
        sheet.evenFooter.left.text = BRAND_SUBTITLE
        sheet.evenFooter.center.text = jurisdiction or "Governed evidence boundary"
        sheet.evenFooter.right.text = f"{fingerprint} · {generated} · Page &P of &N"


def _draw_repeating_pdf_watermark(
    canvas,
    *,
    page_width: float,
    page_height: float,
    sample: bool,
    fingerprint: str,
) -> None:
    canvas.saveState()
    with suppress(AttributeError):
        canvas.setFillAlpha(0.075 if sample else 0.032)
    canvas.setFillColor(colors.HexColor(f"#{_TEAL}"))
    canvas.setFont("Helvetica-Bold", 13 if sample else 12)

    text = "SAMPLE · NOT FOR RESALE" if sample else BRAND_NAME
    for x_fraction in (0.18, 0.5, 0.82):
        for y_fraction in (0.22, 0.5, 0.78):
            canvas.saveState()
            canvas.translate(page_width * x_fraction, page_height * y_fraction)
            canvas.rotate(28)
            canvas.drawCentredString(0, 0, text)
            canvas.setFont("Helvetica", 6)
            canvas.drawCentredString(0, -10, fingerprint)
            canvas.restoreState()
            canvas.setFont("Helvetica-Bold", 13 if sample else 12)
    canvas.restoreState()


def _draw_verification_qr(canvas, *, verification_url: str, page_width: float) -> None:
    qr = QrCodeWidget(verification_url)
    x1, y1, x2, y2 = qr.getBounds()
    size = 17 * mm
    width = x2 - x1
    height = y2 - y1
    drawing = Drawing(
        size,
        size,
        transform=[size / width, 0, 0, size / height, 0, 0],
    )
    drawing.add(qr)
    renderPDF.draw(drawing, canvas, page_width - 31 * mm, 12 * mm)
    canvas.saveState()
    canvas.setFillColor(colors.HexColor(f"#{_DARK_TEAL}"))
    canvas.setFont("Helvetica-Bold", 5.5)
    canvas.drawCentredString(page_width - 22.5 * mm, 10 * mm, "VERIFY RECEIPT")
    canvas.restoreState()


def draw_pdf_branding(
    canvas,
    doc,
    *,
    sample: bool = False,
    order_id: str | None = None,
    jurisdiction: str | None = None,
    generated_at: object | None = None,
    artifact_sha256: str | None = None,
    verification_url: str | None = None,
) -> None:
    """Draw layered institutional watermarking and traceable document controls."""

    generated = _generated_label(generated_at)
    fingerprint = document_fingerprint(
        sample=sample,
        order_id=order_id,
        jurisdiction=jurisdiction,
        generated_at=generated,
        artifact_sha256=artifact_sha256,
    )
    page_width, page_height = doc.pagesize

    _draw_repeating_pdf_watermark(
        canvas,
        page_width=page_width,
        page_height=page_height,
        sample=sample,
        fingerprint=fingerprint,
    )

    canvas.saveState()
    with suppress(AttributeError):
        canvas.setFillAlpha(0.11 if sample else 0.045)
    canvas.setFillColor(colors.HexColor(f"#{_TEAL}"))
    canvas.setFont("Helvetica-Bold", 28 if sample else 24)
    canvas.translate(page_width / 2, page_height / 2)
    canvas.rotate(28)
    canvas.drawCentredString(
        0,
        0,
        "SAMPLE — GAIA FISCAL INTELLIGENCE" if sample else BRAND_NAME,
    )
    canvas.restoreState()

    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#B8CBC6"))
    canvas.setLineWidth(0.6)
    canvas.rect(8 * mm, 8 * mm, page_width - 16 * mm, page_height - 16 * mm, stroke=1, fill=0)
    canvas.restoreState()

    canvas.saveState()
    canvas.setFillColor(colors.HexColor(f"#{_DARK_TEAL}"))
    canvas.rect(0, page_height - 8 * mm, page_width, 8 * mm, stroke=0, fill=1)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 7.5)
    canvas.drawString(10 * mm, page_height - 5.2 * mm, BRAND_NAME)
    classification = (
        "DEMONSTRATION SAMPLE · NOT FOR RESALE" if sample else "GOVERNED CUSTOMER DELIVERABLE"
    )
    canvas.drawCentredString(page_width / 2, page_height - 5.2 * mm, classification)
    canvas.drawRightString(page_width - 10 * mm, page_height - 5.2 * mm, fingerprint)
    canvas.restoreState()

    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#65726F"))
    left = BRAND_NAME
    if jurisdiction:
        left = f"{left} · {jurisdiction}"
    canvas.drawString(12 * mm, 5.5 * mm, left)

    right_parts = [fingerprint]
    if order_id:
        right_parts.append(f"Order {order_id}")
    right_parts.append(generated)
    right_parts.append(f"Page {doc.page}")
    right_margin = page_width - (35 * mm if verification_url and not sample else 12 * mm)
    canvas.drawRightString(right_margin, 5.5 * mm, " · ".join(right_parts))
    canvas.restoreState()

    if verification_url and not sample:
        _draw_verification_qr(
            canvas,
            verification_url=verification_url,
            page_width=page_width,
        )
