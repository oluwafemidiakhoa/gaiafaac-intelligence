from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from gaiafaac_api.database.enums import (
    ProcessingStatus,
    ReportedUnit,
    SourceStatus,
    VerificationStatus,
)
from gaiafaac_api.database.igr_models import IgrPeriodType, StateIgrRecord
from gaiafaac_api.database.models import State
from gaiafaac_api.pipeline.errors import ImportContractError, StateNormalizationError
from gaiafaac_api.pipeline.states import StateNormalizer, normalize_state_key
from gaiafaac_api.services.source_documents import register_source_document

NBS_IGR_ORG = "National Bureau of Statistics (NBS)"
NBS_IGR_2024_SOURCE_URL = (
    "https://microdata.nigerianstat.gov.ng/index.php/catalog/170/download/1294/IGR_2024.zip"
)
_WORKBOOK_MEMBER = "IGR_DATA_2019_2024.xlsx"
_SUPPORTED_YEARS = {2024}
_CENT = Decimal("0.01")
_RECONCILIATION_TOLERANCE = Decimal("0.02")


@dataclass(frozen=True)
class IgrImportResult:
    source_document_id: str
    fiscal_year: int
    records_imported: int
    total_igr: Decimal


def _money(value: object, *, field: str, row_number: int) -> Decimal:
    if value is None:
        raise ImportContractError(f"Missing {field} at workbook row {row_number}")
    try:
        amount = Decimal(str(value)).quantize(_CENT, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as error:
        raise ImportContractError(
            f"Invalid {field} value at workbook row {row_number}: {value!r}"
        ) from error
    if amount < 0:
        raise ImportContractError(f"Negative {field} at workbook row {row_number}: {amount}")
    return amount


def _header_indexes(values: tuple[object, ...]) -> dict[str, int] | None:
    normalized = {
        normalize_state_key(str(value)): index
        for index, value in enumerate(values)
        if value is not None and str(value).strip()
    }
    required = {
        "state": "state",
        "total tax": "total_tax",
        "mdas revenue": "mdas_revenue",
        "total": "total",
    }
    if not set(required).issubset(normalized):
        return None
    return {target: normalized[source] for source, target in required.items()}


def _read_sheet(zip_path: Path, fiscal_year: int):
    try:
        with zipfile.ZipFile(zip_path) as archive:
            bad_member = archive.testzip()
            if bad_member is not None:
                raise ImportContractError(f"Corrupt NBS ZIP member: {bad_member}")
            try:
                workbook_bytes = archive.read(_WORKBOOK_MEMBER)
            except KeyError as error:
                raise ImportContractError(
                    f"NBS archive is missing required member {_WORKBOOK_MEMBER!r}"
                ) from error
    except zipfile.BadZipFile as error:
        raise ImportContractError(f"Source is not a valid ZIP archive: {zip_path}") from error

    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet_name = str(fiscal_year)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ImportContractError(f"NBS workbook has no sheet for fiscal year {fiscal_year}")
    return workbook, workbook[sheet_name]


def import_nbs_igr_zip(
    session: Session,
    *,
    path: Path,
    fiscal_year: int,
    source_url: str = NBS_IGR_2024_SOURCE_URL,
) -> IgrImportResult:
    """Import a verified NBS annual IGR sheet into unpublished review records."""
    if fiscal_year not in _SUPPORTED_YEARS:
        raise ImportContractError(
            f"NBS IGR layout for {fiscal_year} is not verified; supported years: "
            f"{sorted(_SUPPORTED_YEARS)}"
        )

    resolved = path.expanduser().resolve(strict=True)
    if not resolved.is_file():
        raise ImportContractError(f"IGR source is not a regular file: {resolved}")

    if session.scalar(
        select(StateIgrRecord.id).where(
            StateIgrRecord.fiscal_year == fiscal_year,
            StateIgrRecord.period_type == IgrPeriodType.ANNUAL,
            StateIgrRecord.is_demo.is_(False),
        )
    ) is not None:
        raise ImportContractError(f"Non-demo annual IGR records already exist for {fiscal_year}")

    try:
        source = register_source_document(
            session,
            path=resolved,
            source_organization=NBS_IGR_ORG,
            mime_type="application/zip",
            source_url=source_url,
            source_status=SourceStatus.PENDING,
            processing_status=ProcessingStatus.PROCESSING,
            is_demo=False,
            commit=False,
        )

        workbook, sheet = _read_sheet(resolved, fiscal_year)
        try:
            header_indexes: dict[str, int] | None = None
            header_row = 0
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                indexes = _header_indexes(row)
                if indexes is not None:
                    header_indexes = indexes
                    header_row = row_number
                    break
            if header_indexes is None:
                raise ImportContractError(
                    f"Could not find verified IGR headers on sheet {fiscal_year}"
                )

            normalizer = StateNormalizer.from_session(session)
            expected_states = list(session.scalars(select(State).order_by(State.code)))
            expected_ids = {state.id for state in expected_states}
            seen_ids: set = set()
            records: list[StateIgrRecord] = []
            total_igr = Decimal("0.00")

            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                state_value = row[header_indexes["state"]]
                if state_value is None or not str(state_value).strip():
                    continue
                try:
                    match = normalizer.match(str(state_value))
                except StateNormalizationError:
                    # Rows below the jurisdiction table (totals/notes) are ignored only when
                    # they are not numbered data rows. A numbered unknown state is blocking.
                    serial = row[0] if row else None
                    if isinstance(serial, (int, float)) or str(serial).strip().isdigit():
                        raise ImportContractError(
                            f"Unknown IGR jurisdiction at workbook row {row_number}: {state_value!r}"
                        )
                    continue

                if match.state.id in seen_ids:
                    raise ImportContractError(
                        f"Duplicate IGR jurisdiction at workbook row {row_number}: {state_value!r}"
                    )

                total_tax = _money(
                    row[header_indexes["total_tax"]],
                    field="Total Tax",
                    row_number=row_number,
                )
                mdas_revenue = _money(
                    row[header_indexes["mdas_revenue"]],
                    field="MDAs Revenue",
                    row_number=row_number,
                )
                total_value = row[header_indexes["total"]]
                total = _money(total_value, field="Total", row_number=row_number)

                difference = abs((total_tax + mdas_revenue) - total)
                if difference > _RECONCILIATION_TOLERANCE:
                    raise ImportContractError(
                        "NBS IGR row does not reconcile at workbook row "
                        f"{row_number}: Total Tax {total_tax} + MDAs Revenue {mdas_revenue} "
                        f"!= Total {total} (difference {difference})"
                    )

                records.append(
                    StateIgrRecord(
                        state_id=match.state.id,
                        source_document_id=source.id,
                        fiscal_year=fiscal_year,
                        period_type=IgrPeriodType.ANNUAL,
                        quarter=None,
                        period_start=date(fiscal_year, 1, 1),
                        period_end=date(fiscal_year, 12, 31),
                        igr_amount=total,
                        igr_amount_original=str(total_value),
                        reported_unit=ReportedUnit.NAIRA,
                        publication_date=source.publication_date,
                        source_page=None,
                        source_table=sheet.title,
                        verification_status=VerificationStatus.REQUIRES_REVIEW,
                        is_demo=False,
                        is_published=False,
                    )
                )
                seen_ids.add(match.state.id)
                total_igr += total

            missing = expected_ids - seen_ids
            extra = seen_ids - expected_ids
            if missing or extra or len(records) != 37:
                missing_codes = sorted(state.code for state in expected_states if state.id in missing)
                raise ImportContractError(
                    "NBS IGR jurisdiction coverage failed: "
                    f"records={len(records)}, missing={missing_codes}, extra_count={len(extra)}"
                )

            session.add_all(records)
            source.processing_status = ProcessingStatus.READY_FOR_REVIEW
            source.source_status = SourceStatus.READY_FOR_REVIEW
            session.commit()
        finally:
            workbook.close()
    except Exception:
        session.rollback()
        raise

    return IgrImportResult(
        source_document_id=str(source.id),
        fiscal_year=fiscal_year,
        records_imported=len(records),
        total_igr=total_igr.quantize(_CENT),
    )
