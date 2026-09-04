from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from gaiafaac_api.pipeline.extraction.oagf_lga_table_iv import (
    ExtractedLgaAllocation,
    ExtractedLgaTable,
)

_EXPECTED_JURISDICTIONS = 774
_SKIP_NAMES = {"", "local government councils", "local government council", "total", "grand total"}


@dataclass(frozen=True)
class _Panel:
    state: int
    lga: int
    net_statutory: int | None
    deduction: int | None
    ecology_share: int | None
    ecology_transfer: int | None
    net_ecology: int | None
    vat: int | None
    total_net: int


def _clean(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(Decimal(str(value)), "f")
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _key(value: object | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean(value).casefold()).strip()


def _money(value: object | None) -> Decimal | None:
    text = _clean(value)
    if not text or text in {"-", "–", "—"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    normalized = text.strip("()₦N ").replace(",", "")
    try:
        amount = Decimal(normalized)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def _is_total_name(value: str) -> bool:
    folded = value.casefold().strip()
    return folded in _SKIP_NAMES or folded.endswith(" total")


def _column_labels(rows: list[tuple[Any, ...]], header_limit: int) -> list[str]:
    width = max((len(row) for row in rows[:header_limit]), default=0)
    parts: list[list[str]] = [[] for _ in range(width)]
    for row in rows[:header_limit]:
        for index in range(width):
            value = _clean(row[index]) if index < len(row) else ""
            if value:
                parts[index].append(value)
    return [_key(" ".join(values)) for values in parts]


def _find_index(labels: list[str], start: int, end: int, *needles: str) -> int | None:
    for index in range(start, end):
        label = labels[index]
        if all(needle in label for needle in needles):
            return index
    return None


def _resolve_panels(labels: list[str]) -> list[_Panel]:
    lga_columns = [
        index
        for index, label in enumerate(labels)
        if "local government" in label or "area council" in label
    ]
    panels: list[_Panel] = []
    for panel_number, lga_index in enumerate(lga_columns):
        start = 0 if panel_number == 0 else lga_columns[panel_number - 1] + 1
        end = lga_columns[panel_number + 1] if panel_number + 1 < len(lga_columns) else len(labels)
        state_index = next(
            (
                index
                for index in range(lga_index - 1, start - 1, -1)
                if labels[index] == "state" or labels[index].startswith("state ")
            ),
            None,
        )
        total_net = _find_index(labels, lga_index + 1, end, "total", "net", "allocation")
        if state_index is None or total_net is None:
            continue
        panels.append(
            _Panel(
                state=state_index,
                lga=lga_index,
                net_statutory=_find_index(labels, lga_index + 1, end, "net", "statutory"),
                deduction=_find_index(labels, lga_index + 1, end, "deduction"),
                ecology_share=_find_index(labels, lga_index + 1, end, "total", "share", "ecology"),
                ecology_transfer=_find_index(labels, lga_index + 1, end, "transfer", "ecology"),
                net_ecology=_find_index(labels, lga_index + 1, end, "net", "share", "ecology"),
                vat=(
                    _find_index(labels, lga_index + 1, end, "value", "added", "tax")
                    or _find_index(labels, lga_index + 1, end, "vat")
                ),
                total_net=total_net,
            )
        )
    return panels


def _value(row: tuple[Any, ...], index: int | None) -> object | None:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _header_anchor(rows: list[tuple[Any, ...]]) -> int | None:
    for row_index, row in enumerate(rows[:30]):
        if any("local government" in _key(value) or "area council" in _key(value) for value in row):
            return row_index
    return None


def extract_oagf_table_iv_excel(path: Path) -> ExtractedLgaTable:
    """Extract governed OAGF Table IV LGA rows from an official Excel workbook.

    The workbook is treated as evidence, not as trusted application input. The
    extractor discovers Table IV-style columns, carries merged/blank state names
    down rows, deduplicates jurisdictions, and fails closed unless exactly 774
    Nigerian LGA/Area Council rows are observed.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    extracted: list[ExtractedLgaAllocation] = []
    warnings: list[str] = []
    seen: set[tuple[str, str]] = set()
    duplicate_count = 0
    resolved_sheets = 0

    try:
        for worksheet in workbook.worksheets:
            rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
            anchor = _header_anchor(rows)
            if anchor is None:
                continue

            labels = _column_labels(rows, min(len(rows), anchor + 6))
            panels = _resolve_panels(labels)
            if not panels:
                warnings.append(
                    f"Excel sheet {worksheet.title!r} contains an LGA header but its Table IV columns could not be resolved."
                )
                continue

            resolved_sheets += 1
            current_states = ["" for _ in panels]
            for _source_row, row in enumerate(rows[anchor + 1 :], start=anchor + 2):
                for panel_number, panel in enumerate(panels):
                    observed_state = _clean(_value(row, panel.state))
                    state_key = _key(observed_state)
                    if (
                        observed_state
                        and state_key not in {"state", "states"}
                        and not observed_state.casefold().endswith(" total")
                        and re.search(r"[A-Za-z]", observed_state)
                    ):
                        current_states[panel_number] = observed_state

                    state = current_states[panel_number]
                    lga = _clean(_value(row, panel.lga))
                    total_original = _clean(_value(row, panel.total_net))
                    total_net = _money(_value(row, panel.total_net))
                    if (
                        not state
                        or _is_total_name(lga)
                        or total_net is None
                        or not re.search(r"[A-Za-z]", lga)
                    ):
                        continue

                    key = (_key(state), _key(lga))
                    if key in seen:
                        duplicate_count += 1
                        continue
                    seen.add(key)

                    originals = {
                        "net_statutory_allocation": _clean(_value(row, panel.net_statutory))
                        or None,
                        "deduction_amount": _clean(_value(row, panel.deduction)) or None,
                        "ecology_share": _clean(_value(row, panel.ecology_share)) or None,
                        "ecology_transfer": _clean(_value(row, panel.ecology_transfer)) or None,
                        "net_ecology_share": _clean(_value(row, panel.net_ecology)) or None,
                        "vat_amount": _clean(_value(row, panel.vat)) or None,
                        "total_net_allocation": total_original,
                    }
                    extracted.append(
                        ExtractedLgaAllocation(
                            state_name=state,
                            local_government_name=lga,
                            net_statutory_allocation=_money(_value(row, panel.net_statutory)),
                            deduction_amount=_money(_value(row, panel.deduction)),
                            ecology_share=_money(_value(row, panel.ecology_share)),
                            ecology_transfer=_money(_value(row, panel.ecology_transfer)),
                            net_ecology_share=_money(_value(row, panel.net_ecology)),
                            vat_amount=_money(_value(row, panel.vat)),
                            total_net_allocation=total_net,
                            originals=originals,
                            page=None,  # Excel provenance is row-based rather than page-based.
                        )
                    )

        if resolved_sheets == 0:
            warnings.append("No OAGF Table IV worksheet could be resolved in the Excel source.")
        if duplicate_count:
            warnings.append(
                f"OAGF Table IV Excel source contained {duplicate_count} duplicate jurisdiction row(s)."
            )
        if len(extracted) != _EXPECTED_JURISDICTIONS:
            warnings.append(
                "OAGF Table IV Excel coverage is incomplete or ambiguous: "
                f"expected {_EXPECTED_JURISDICTIONS} jurisdictions, extracted {len(extracted)}."
            )

        return ExtractedLgaTable(
            rows=extracted,
            warnings=warnings,
            requires_review=bool(warnings),
        )
    finally:
        workbook.close()
