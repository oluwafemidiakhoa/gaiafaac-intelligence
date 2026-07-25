from __future__ import annotations

from pathlib import Path

from gaiafaac_api.pipeline.extraction.schema import (
    ALLOCATION_FIELDS,
    CellProvenance,
    ExtractedAllocationRow,
    ExtractedAllocationTable,
)

_EXTRA_COLUMNS = ("extraction_confidence", "data_label")
_XLSX_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


def _cell_text(value: object) -> tuple[str | None, bool]:
    """Return (original_text, is_float). Floats are risky for exact money."""
    if value is None:
        return None, False
    if isinstance(value, float):
        return str(value), True
    return str(value), False


class GenericExcelAdapter:
    """Read the first sheet of an XLSX into the normalized schema.

    Excel stores numbers as floats, which cannot represent exact kobo reliably.
    Any monetary cell that arrives as a float is preserved verbatim but flags the
    table for human review rather than being silently trusted.
    """

    name = "generic_excel"

    def supports(self, path: Path, mime_type: str) -> bool:
        return path.suffix.casefold() in {".xlsx", ".xlsm"} or mime_type in _XLSX_MIME

    def extract(self, path: Path) -> ExtractedAllocationTable:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            row_iter = worksheet.iter_rows(values_only=True)
            header_values = next(row_iter, None) or ()
            index = {
                str(name).strip(): position
                for position, name in enumerate(header_values)
                if name is not None
            }

            rows: list[ExtractedAllocationRow] = []
            warnings: list[str] = []
            requires_review = False

            for source_row, values in enumerate(row_iter, start=2):
                if values is None or all(value is None for value in values):
                    continue
                cells: dict[str, CellProvenance] = {}
                for field_name in (*ALLOCATION_FIELDS, *_EXTRA_COLUMNS):
                    position = index.get(field_name)
                    if position is None:
                        continue
                    raw = values[position] if position < len(values) else None
                    text, is_float = _cell_text(raw)
                    if is_float and field_name in ALLOCATION_FIELDS:
                        requires_review = True
                    cells[field_name] = CellProvenance(
                        original_text=text, row=source_row, column=field_name
                    )
                state_position = index.get("state")
                state_text = (
                    _cell_text(values[state_position])[0]
                    if state_position is not None and state_position < len(values)
                    else None
                )
                unit_position = index.get("reported_unit")
                unit_text = (
                    _cell_text(values[unit_position])[0]
                    if unit_position is not None and unit_position < len(values)
                    else None
                )
                rows.append(
                    ExtractedAllocationRow(
                        submitted_state=state_text or "",
                        reported_unit=unit_text,
                        cells=cells,
                        source_row=source_row,
                    )
                )

            if requires_review:
                warnings.append(
                    "One or more monetary cells were stored as Excel numbers (float); "
                    "verify exact values against the source."
                )
            return ExtractedAllocationTable(
                source_organization="",
                adapter_name=self.name,
                rows=rows,
                warnings=warnings,
                requires_review=requires_review,
            )
        finally:
            workbook.close()
