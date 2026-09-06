from __future__ import annotations

import io

from openpyxl import load_workbook

from gaiafaac_api.services.one_time_exports import build_one_time_excel, build_one_time_pdf


def _artifact() -> dict:
    return {
        "schema": "gaia-one-time-historical-export-v1",
        "captured_at": "2026-09-05T23:00:00+00:00",
        "request": {
            "state_slug": "lagos",
            "state_code": "LA",
            "domain": "igr",
            "start_year": 2023,
            "end_year": 2024,
        },
        "rows": [
            {
                "domain": "igr",
                "state_name": "Lagos",
                "state_code": "LA",
                "period": "2024",
                "value": "1261556415048.56",
                "unit": "NGN",
                "source_organization": "National Bureau of Statistics",
                "source_url": "https://example.gov.ng/nbs-igr.xlsx",
                "source_sha256": "a" * 64,
                "verification_status": "human_verified",
            }
        ],
    }


def test_one_time_excel_export_contains_summary_and_evidence_sheet():
    filename, media_type, body = build_one_time_excel(
        purchase_id="12345678-1234-5678-1234-567812345678",
        product_code="historical_evidence_export",
        amount_naira="75000",
        currency="NGN",
        completed_at="2026-09-05T23:30:00+00:00",
        artifact=_artifact(),
    )

    assert filename.endswith(".xlsx")
    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(io.BytesIO(body), data_only=True)
    assert "Summary" in workbook.sheetnames
    assert "Evidence Rows" in workbook.sheetnames
    assert workbook["Summary"]["A1"].value == "Gaia Fiscal Intelligence"
    assert workbook["Evidence Rows"]["A2"].value == "igr"
    assert workbook["Evidence Rows"]["H2"].value == "https://example.gov.ng/nbs-igr.xlsx"


def test_one_time_pdf_export_is_downloadable_pdf():
    filename, media_type, body = build_one_time_pdf(
        purchase_id="12345678-1234-5678-1234-567812345678",
        product_code="historical_evidence_export",
        amount_naira="75000",
        currency="NGN",
        completed_at="2026-09-05T23:30:00+00:00",
        artifact=_artifact(),
    )

    assert filename.endswith(".pdf")
    assert media_type == "application/pdf"
    assert body.startswith(b"%PDF-")
    assert len(body) > 1000
