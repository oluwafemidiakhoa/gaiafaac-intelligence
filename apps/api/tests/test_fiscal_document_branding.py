from __future__ import annotations

import io

from openpyxl import load_workbook

from gaiafaac_api.services.branded_one_time_exports import (
    build_one_time_excel,
    build_one_time_pdf,
)
from gaiafaac_api.services.document_branding import (
    BRAND_NAME,
    SAMPLE_NOTICE,
    document_fingerprint,
)
from gaiafaac_api.services.project_receipts import canonical_artifact_sha256


def _artifact() -> dict:
    return {
        "schema": "gaia-one-time-historical-export-v1",
        "captured_at": "2026-09-06T13:00:00+00:00",
        "request": {
            "state_slug": "lagos",
            "state_code": "LA",
            "domain": "igr",
            "start_year": 2023,
            "end_year": 2024,
        },
        "rows": [
            {
                "domain": "igr",
                "state_name": "Lagos",
                "state_code": "LA",
                "period": "2024",
                "value": "1261556415048.56",
                "unit": "NGN",
                "source_organization": "National Bureau of Statistics",
                "source_url": "https://example.gov.ng/nbs-igr.xlsx",
                "source_sha256": "a" * 64,
                "verification_status": "human_verified",
            }
        ],
    }


def test_paid_excel_uses_document_control_and_fingerprint_on_every_sheet():
    purchase_id = "12345678-1234-5678-1234-567812345678"
    artifact = _artifact()
    artifact_sha256 = canonical_artifact_sha256(artifact)
    filename, media_type, body = build_one_time_excel(
        purchase_id=purchase_id,
        product_code="historical_evidence_export",
        amount_naira="75000",
        currency="NGN",
        completed_at="2026-09-06T13:30:00+00:00",
        artifact=artifact,
        jurisdiction="Lagos",
    )

    assert filename.startswith("gaia-fiscal-intelligence-")
    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(io.BytesIO(body), data_only=True)
    assert workbook["Summary"]["A1"].value == BRAND_NAME
    assert "Document Control" in workbook.sheetnames

    fingerprint = document_fingerprint(
        sample=False,
        order_id=purchase_id,
        jurisdiction="Lagos",
        generated_at=artifact["captured_at"],
        artifact_sha256=artifact_sha256,
    )
    control_values = [
        workbook["Document Control"].cell(row=row, column=2).value
        for row in range(1, workbook["Document Control"].max_row + 1)
    ]
    assert fingerprint in control_values
    assert artifact_sha256 in control_values
    assert any("/verify/project/" in str(value) for value in control_values if value)
    assert any(
        "Customer pays for governed fiscal intelligence" in str(value)
        for value in control_values
        if value
    )
    for sheet in workbook.worksheets:
        assert BRAND_NAME in (sheet.oddHeader.left.text or "")
        assert fingerprint in (sheet.oddHeader.center.text or "")
        assert fingerprint in (sheet.oddFooter.right.text or "")


def test_sample_excel_is_visibly_classified_and_traceable():
    artifact = _artifact()
    artifact_sha256 = canonical_artifact_sha256(artifact)
    filename, _media_type, body = build_one_time_excel(
        purchase_id="SAMPLE-lagos-2026",
        product_code="decision_pack",
        amount_naira="50000",
        currency="NGN",
        completed_at="Not applicable — demonstration sample",
        artifact=artifact,
        sample=True,
        jurisdiction="Lagos",
    )

    assert filename == "gaia-fiscal-intelligence-sample-decision-pack.xlsx"
    workbook = load_workbook(io.BytesIO(body), data_only=True)
    assert workbook["Summary"]["A3"].value == SAMPLE_NOTICE
    assert "Document Control" in workbook.sheetnames
    fingerprint = document_fingerprint(
        sample=True,
        order_id=None,
        jurisdiction="Lagos",
        generated_at=artifact["captured_at"],
        artifact_sha256=artifact_sha256,
    )
    for sheet in workbook.worksheets:
        assert "SAMPLE" in (sheet.oddHeader.right.text or "")
        assert fingerprint in (sheet.oddHeader.center.text or "")


def test_paid_and_sample_pdf_are_real_branded_documents():
    paid_filename, paid_media_type, paid_body = build_one_time_pdf(
        purchase_id="12345678-1234-5678-1234-567812345678",
        product_code="decision_pack",
        amount_naira="50000",
        currency="NGN",
        completed_at="2026-09-06T13:30:00+00:00",
        artifact=_artifact(),
        jurisdiction="Lagos",
    )
    sample_filename, sample_media_type, sample_body = build_one_time_pdf(
        purchase_id="SAMPLE-lagos-2026",
        product_code="decision_pack",
        amount_naira="50000",
        currency="NGN",
        completed_at="Not applicable — demonstration sample",
        artifact=_artifact(),
        sample=True,
        jurisdiction="Lagos",
    )

    assert paid_filename.startswith("gaia-fiscal-intelligence-decision-pack-")
    assert sample_filename == "gaia-fiscal-intelligence-sample-decision-pack.pdf"
    assert paid_media_type == sample_media_type == "application/pdf"
    assert paid_body.startswith(b"%PDF-")
    assert sample_body.startswith(b"%PDF-")
    assert len(paid_body) > 1000
    assert len(sample_body) > 1000
